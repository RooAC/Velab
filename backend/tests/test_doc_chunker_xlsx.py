"""
doc_chunker.py 中 .xlsx 提取的单元测试
"""

from __future__ import annotations

from pathlib import Path

import pytest

openpyxl = pytest.importorskip("openpyxl")

from services.doc_chunker import DocumentChunker


def _make_xlsx(tmp_path: Path) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Diagnose"
    ws1.append(["ECU", "Code", "Action"])
    ws1.append(["iCGM", "0x1234", "重启网关"])
    ws1.append(["MPU", "0x5678", "刷写 bootloader"])
    ws2 = wb.create_sheet("Notes")
    ws2.append(["规范条款", "说明"])
    ws2.append(["状态机", "INIT → DOWNLOAD → VERIFY"])

    path = tmp_path / "spec.xlsx"
    wb.save(path)
    return path


class TestXlsxExtraction:
    def test_extracts_text_across_sheets(self, tmp_path):
        path = _make_xlsx(tmp_path)
        text = DocumentChunker._extract_xlsx_text(path)
        assert "Sheet: Diagnose" in text
        assert "Sheet: Notes" in text
        assert "iCGM" in text and "0x1234" in text
        assert "重启网关" in text
        assert "INIT → DOWNLOAD → VERIFY" in text

    def test_chunk_file_supports_xlsx(self, tmp_path):
        path = _make_xlsx(tmp_path)
        chunks = DocumentChunker(chunk_size=200).chunk_file(path)
        assert chunks, "至少应有一个 chunk"
        joined = "\n".join(c.content for c in chunks)
        assert "iCGM" in joined
        assert "状态机" in joined

    def test_returns_empty_on_corrupt_xlsx(self, tmp_path):
        path = tmp_path / "broken.xlsx"
        path.write_bytes(b"not a real xlsx")
        text = DocumentChunker._extract_xlsx_text(path)
        assert text == ""

    def test_skips_empty_rows(self, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["Header"])
        ws.append([None, None, None])  # 整行空
        ws.append(["Value"])
        path = tmp_path / "sparse.xlsx"
        wb.save(path)

        text = DocumentChunker._extract_xlsx_text(path)
        lines = [ln for ln in text.split("\n") if ln and not ln.startswith("## Sheet")]
        # 仅两行非空数据，不包括 None 行
        assert lines == ["Header", "Value"]
